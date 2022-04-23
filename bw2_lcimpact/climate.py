import os
from collections import defaultdict
from openpyxl import load_workbook

from .base import LCIA, data_dir

CATEGORIES = {
    ("air",),
    ("air", "non-urban air or from high stacks"),
    ("air", "lower stratosphere + upper troposphere"),
    ("air", "low population density, long-term"),
    ("air", "urban air close to ground"),
}


def get_values_by_column(column):
    wb = load_workbook(os.path.join(data_dir, "climate-change.xlsx"), data_only=True, read_only=True)
    mapping = defaultdict(list)
    for key, value in wb['mapping'].iter_rows(min_col=1, max_col=2, values_only=True):
        if key:
            mapping[key].append(value)

    labels = [val[0] for val in wb['cfs'].iter_rows(min_col=1, max_col=1, min_row=2, values_only=True) if val is not None]
    values = [val[0] for val in wb['cfs'].iter_rows(min_col=column, max_col=column, min_row=2, values_only=True) if val is not None]

    cfs = {}
    for label, value in zip(labels, values):
        for mapped in mapping.get(label, []):
            cfs[mapped] = value

    return cfs


class ClimateChange(LCIA):
    geocollection = None

    def setup_geocollections(self):
        return

    def regional_cfs(self):
        for obj in self.global_cfs():
            yield obj

    def global_cfs(self):
        cfs = get_values_by_column(self.column + 1)

        for flow in self.db:
            if flow["name"] in cfs and tuple(flow["categories"]) in CATEGORIES:
                yield ((flow.key, cfs[flow["name"]], "GLO"))


class ClimateChangeHumanHealth(ClimateChange):
    unit = "DALY/kg"
    description = """"""
    url = "http://lc-impact.eu/human-health-climate-change"


class ClimateChangeHumanHealthCertain100Years(ClimateChangeHumanHealth):
    name = (
        "LC-IMPACT",
        "Climate Change",
        "Human Health",
        "Marginal",
        "Certain",
        "100 Years",
    )
    column = 1


class ClimateChangeHumanHealthAll100Years(ClimateChangeHumanHealth):
    name = (
        "LC-IMPACT",
        "Climate Change",
        "Human Health",
        "Marginal",
        "All",
        "100 Years",
    )
    column = 2


class ClimateChangeHumanHealthCertainInfinite(ClimateChangeHumanHealth):
    name = (
        "LC-IMPACT",
        "Climate Change",
        "Human Health",
        "Marginal",
        "Certain",
        "Infinite",
    )
    column = 3


class ClimateChangeHumanHealthAllInfinite(ClimateChangeHumanHealth):
    name = (
        "LC-IMPACT",
        "Climate Change",
        "Human Health",
        "Marginal",
        "All",
        "Infinite",
    )
    column = 4


class ClimateChangeTerrestrialEcosystems(ClimateChange):
    unit = "PDF year/kg"
    description = """"""
    url = "http://lc-impact.eu/ecosystem-quality-climate-change"


class ClimateChangeTerrestrialEcosystemsCertain100Years(
    ClimateChangeTerrestrialEcosystems
):
    name = (
        "LC-IMPACT",
        "Climate Change",
        "Terrestrial Ecosystems",
        "Marginal",
        "Certain",
        "100 Years",
    )
    column = 5


class ClimateChangeTerrestrialEcosystemsAll100Years(ClimateChangeTerrestrialEcosystems):
    name = (
        "LC-IMPACT",
        "Climate Change",
        "Terrestrial Ecosystems",
        "Marginal",
        "All",
        "100 Years",
    )
    column = 6


class ClimateChangeTerrestrialEcosystemsCertainInfinite(
    ClimateChangeTerrestrialEcosystems
):
    name = (
        "LC-IMPACT",
        "Climate Change",
        "Terrestrial Ecosystems",
        "Marginal",
        "Certain",
        "Infinite",
    )
    column = 7


class ClimateChangeTerrestrialEcosystemsAllInfinite(ClimateChangeTerrestrialEcosystems):
    name = (
        "LC-IMPACT",
        "Climate Change",
        "Terrestrial Ecosystems",
        "Marginal",
        "All",
        "Infinite",
    )
    column = 8


class ClimateChangeAquaticEcosystems(ClimateChangeTerrestrialEcosystems):
    description = """"""


class ClimateChangeAquaticEcosystemsAll100Years(ClimateChangeAquaticEcosystems):
    name = (
        "LC-IMPACT",
        "Climate Change",
        "Aquatic Ecosystems",
        "Marginal",
        "All",
        "100 Years",
    )
    column = 10


class ClimateChangeAquaticEcosystemsAllInfinite(ClimateChangeAquaticEcosystems):
    name = (
        "LC-IMPACT",
        "Climate Change",
        "Aquatic Ecosystems",
        "Marginal",
        "All",
        "Infinite",
    )
    column = 12
